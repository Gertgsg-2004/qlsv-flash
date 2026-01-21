# admin_routes.py
from __future__ import annotations

import os
import re
import uuid
import secrets
import string
from io import BytesIO
from datetime import datetime

from flask import (
    Blueprint, render_template, redirect, url_for, flash, abort,
    request, current_app, send_file
)
from flask_login import login_required
from flask_wtf.csrf import validate_csrf
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook

from extensions import db, bcrypt
from models import Class, Exam, Grade, Lesson, Student, Subject
from auth import roles_required


admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

# =========================================================
# Helpers
# =========================================================
ALLOWED_EXTS = {".png", ".jpg", ".jpeg", ".webp"}
EXPORT_HEADERS = ["Mã SV", "Tên", "Email", "Mật khẩu", "Giới tính", "Địa chỉ"]


def _upload_dir() -> str:
    path = os.path.join(current_app.root_path, "static", "uploads")
    os.makedirs(path, exist_ok=True)
    return path


def _norm_cell(x) -> str:
    if x is None:
        return ""
    return str(x).strip()


def _random_password(length: int = 10) -> str:
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))


def _csrf_or_flash(redirect_endpoint: str, **kwargs):
    """Validate CSRF; nếu lỗi thì flash + redirect."""
    try:
        validate_csrf(request.form.get("csrf_token", ""))
        return None
    except Exception:
        flash("CSRF token không hợp lệ.", "danger")
        return redirect(url_for(redirect_endpoint, **kwargs))


# =========================================================
# HOME (Admin)
# =========================================================
@admin_bp.get("/")
@login_required
@roles_required("admin")
def admin_home():
    return render_template("admin/home.html")


# =========================================================
# USERS + ROLE
# =========================================================
@admin_bp.get("/users")
@login_required
@roles_required("admin")
def admin_users():
    users = Student.query.order_by(Student.id.asc()).all()
    return render_template("admin/users.html", users=users)


@admin_bp.post("/users/<int:user_id>/role")
@login_required
@roles_required("admin")
def admin_update_role(user_id: int):
    user = Student.query.get(user_id)
    if not user:
        abort(404)

    bad = _csrf_or_flash("admin.admin_users")
    if bad:
        return bad

    new_role = (request.form.get("role") or "").strip()
    if new_role not in ("admin", "manager", "student"):
        flash("Role không hợp lệ.", "danger")
        return redirect(url_for("admin.admin_users"))

    user.role = new_role
    db.session.commit()
    flash(f"Đã cập nhật role cho {user.email} => {new_role}", "success")
    return redirect(url_for("admin.admin_users"))


# =========================================================
# SUBJECTS
# =========================================================
@admin_bp.get("/subjects")
@login_required
@roles_required("admin")
def admin_subjects():
    q = (request.args.get("q") or "").strip()
    query = Subject.query
    if q:
        query = query.filter(
            (Subject.code.ilike(f"%{q}%")) |
            (Subject.name.ilike(f"%{q}%"))
        )
    subjects = query.order_by(Subject.id.desc()).all()
    return render_template("admin/subjects.html", subjects=subjects, q=q)


@admin_bp.route("/subjects/new", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def admin_subject_create():
    if request.method == "GET":
        return render_template("admin/subject_form.html", mode="create", subject=None)

    bad = _csrf_or_flash("admin.admin_subject_create")
    if bad:
        return bad

    code = (request.form.get("code") or "").strip()
    name = (request.form.get("name") or "").strip()
    credits_raw = (request.form.get("credits") or "3").strip()
    description = (request.form.get("description") or "").strip()

    if not code or not name:
        flash("Vui lòng nhập Code và Tên môn học.", "danger")
        return redirect(url_for("admin.admin_subject_create"))

    try:
        credits = int(credits_raw)
        if credits <= 0:
            raise ValueError()
    except Exception:
        flash("Tín chỉ phải là số nguyên dương.", "danger")
        return redirect(url_for("admin.admin_subject_create"))

    if Subject.query.filter_by(code=code).first():
        flash("Code môn học đã tồn tại.", "danger")
        return redirect(url_for("admin.admin_subject_create"))

    s = Subject(code=code, name=name, credits=credits, description=description or None)
    db.session.add(s)
    db.session.commit()
    flash("Đã thêm môn học!", "success")
    return redirect(url_for("admin.admin_subjects"))


@admin_bp.route("/subjects/<int:subject_id>/edit", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def admin_subject_edit(subject_id: int):
    subject = Subject.query.get(subject_id)
    if not subject:
        abort(404)

    if request.method == "GET":
        return render_template("admin/subject_form.html", mode="edit", subject=subject)

    bad = _csrf_or_flash("admin.admin_subject_edit", subject_id=subject_id)
    if bad:
        return bad

    code = (request.form.get("code") or "").strip()
    name = (request.form.get("name") or "").strip()
    credits_raw = (request.form.get("credits") or "3").strip()
    description = (request.form.get("description") or "").strip()

    if not code or not name:
        flash("Vui lòng nhập Code và Tên môn học.", "danger")
        return redirect(url_for("admin.admin_subject_edit", subject_id=subject_id))

    try:
        credits = int(credits_raw)
        if credits <= 0:
            raise ValueError()
    except Exception:
        flash("Tín chỉ phải là số nguyên dương.", "danger")
        return redirect(url_for("admin.admin_subject_edit", subject_id=subject_id))

    exist = Subject.query.filter(Subject.code == code, Subject.id != subject.id).first()
    if exist:
        flash("Code môn học đã tồn tại.", "danger")
        return redirect(url_for("admin.admin_subject_edit", subject_id=subject_id))

    subject.code = code
    subject.name = name
    subject.credits = credits
    subject.description = description or None
    db.session.commit()

    flash("Đã cập nhật môn học!", "success")
    return redirect(url_for("admin.admin_subjects"))


@admin_bp.post("/subjects/<int:subject_id>/delete")
@login_required
@roles_required("admin")
def admin_subject_delete(subject_id: int):
    subject = Subject.query.get(subject_id)
    if not subject:
        abort(404)

    bad = _csrf_or_flash("admin.admin_subjects")
    if bad:
        return bad

    db.session.delete(subject)
    db.session.commit()
    flash("Đã xoá môn học!", "success")
    return redirect(url_for("admin.admin_subjects"))


# =========================================================
# CLASSES
# =========================================================
@admin_bp.get("/classes")
@login_required
@roles_required("admin")
def admin_classes():
    q = (request.args.get("q") or "").strip()
    query = Class.query
    if q:
        query = query.filter(
            (Class.code.ilike(f"%{q}%")) |
            (Class.name.ilike(f"%{q}%"))
        )
    classes = query.order_by(Class.id.desc()).all()
    subjects = Subject.query.order_by(Subject.name.asc()).all()
    return render_template("admin/classes.html", classes=classes, q=q, subjects=subjects)


@admin_bp.route("/classes/new", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def admin_class_create():
    if request.method == "GET":
        subjects = Subject.query.order_by(Subject.name.asc()).all()
        return render_template("admin/class_form.html", mode="create", c=None, subjects=subjects)

    bad = _csrf_or_flash("admin.admin_class_create")
    if bad:
        return bad

    code = (request.form.get("code") or "").strip()
    name = (request.form.get("name") or "").strip()
    description = (request.form.get("description") or "").strip()

    subject_ids = request.form.getlist("subject_ids")
    subject_ids = [int(x) for x in subject_ids if str(x).isdigit()]

    if not code or not name:
        flash("Vui lòng nhập Code và Tên lớp.", "danger")
        return redirect(url_for("admin.admin_class_create"))

    if Class.query.filter_by(code=code).first():
        flash("Code lớp đã tồn tại.", "danger")
        return redirect(url_for("admin.admin_class_create"))

    c = Class(code=code, name=name, description=description or None)
    if subject_ids:
        c.subjects = Subject.query.filter(Subject.id.in_(subject_ids)).all()

    db.session.add(c)
    db.session.commit()
    flash("Đã thêm lớp!", "success")
    return redirect(url_for("admin.admin_classes"))


@admin_bp.route("/classes/<int:class_id>/edit", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def admin_class_edit(class_id: int):
    c = Class.query.get(class_id)
    if not c:
        abort(404)

    if request.method == "GET":
        subjects = Subject.query.order_by(Subject.name.asc()).all()
        selected_ids = {s.id for s in c.subjects}
        return render_template("admin/class_form.html", mode="edit", c=c, subjects=subjects, selected_ids=selected_ids)

    bad = _csrf_or_flash("admin.admin_class_edit", class_id=class_id)
    if bad:
        return bad

    code = (request.form.get("code") or "").strip()
    name = (request.form.get("name") or "").strip()
    description = (request.form.get("description") or "").strip()

    subject_ids = request.form.getlist("subject_ids")
    subject_ids = [int(x) for x in subject_ids if str(x).isdigit()]

    if not code or not name:
        flash("Vui lòng nhập Code và Tên lớp.", "danger")
        return redirect(url_for("admin.admin_class_edit", class_id=class_id))

    exist = Class.query.filter(Class.code == code, Class.id != c.id).first()
    if exist:
        flash("Code lớp đã tồn tại.", "danger")
        return redirect(url_for("admin.admin_class_edit", class_id=class_id))

    c.code = code
    c.name = name
    c.description = description or None
    c.subjects = Subject.query.filter(Subject.id.in_(subject_ids)).all() if subject_ids else []

    db.session.commit()
    flash("Đã cập nhật lớp!", "success")
    return redirect(url_for("admin.admin_classes"))


@admin_bp.post("/classes/<int:class_id>/delete")
@login_required
@roles_required("admin")
def admin_class_delete(class_id: int):
    c = Class.query.get(class_id)
    if not c:
        abort(404)

    bad = _csrf_or_flash("admin.admin_classes")
    if bad:
        return bad

    db.session.delete(c)
    db.session.commit()
    flash("Đã xoá lớp!", "success")
    return redirect(url_for("admin.admin_classes"))


# =========================================================
# EXAMS
# =========================================================
@admin_bp.get("/exams")
@login_required
@roles_required("admin")
def admin_exams():
    q = (request.args.get("q") or "").strip()
    query = Exam.query.join(Subject)
    if q:
        query = query.filter(
            (Exam.name.ilike(f"%{q}%")) |
            (Subject.name.ilike(f"%{q}%")) |
            (Subject.code.ilike(f"%{q}%"))
        )
    exams = query.order_by(Exam.id.desc()).all()
    subjects = Subject.query.order_by(Subject.name.asc()).all()
    return render_template("admin/exams.html", exams=exams, subjects=subjects, q=q)


@admin_bp.route("/exams/new", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def admin_exam_create():
    subjects = Subject.query.order_by(Subject.name.asc()).all()
    if request.method == "GET":
        return render_template("admin/exam_form.html", mode="create", e=None, subjects=subjects)

    bad = _csrf_or_flash("admin.admin_exam_create")
    if bad:
        return bad

    name = (request.form.get("name") or "").strip()
    subject_id = (request.form.get("subject_id") or "").strip()
    exam_date = (request.form.get("exam_date") or "").strip()
    description = (request.form.get("description") or "").strip()

    if not name or not subject_id.isdigit():
        flash("Vui lòng nhập Tên kỳ thi và chọn Môn học.", "danger")
        return redirect(url_for("admin.admin_exam_create"))

    subject = Subject.query.get(int(subject_id))
    if not subject:
        flash("Môn học không tồn tại.", "danger")
        return redirect(url_for("admin.admin_exam_create"))

    dt = None
    if exam_date:
        try:
            dt = datetime.strptime(exam_date, "%Y-%m-%d").date()
        except Exception:
            flash("Ngày thi không đúng định dạng (YYYY-MM-DD).", "danger")
            return redirect(url_for("admin.admin_exam_create"))

    e = Exam(name=name, subject_id=subject.id, exam_date=dt, description=description or None)
    db.session.add(e)
    db.session.commit()
    flash("Đã thêm kỳ thi!", "success")
    return redirect(url_for("admin.admin_exams"))


@admin_bp.route("/exams/<int:exam_id>/edit", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def admin_exam_edit(exam_id: int):
    e = Exam.query.get(exam_id)
    if not e:
        abort(404)

    subjects = Subject.query.order_by(Subject.name.asc()).all()
    if request.method == "GET":
        return render_template("admin/exam_form.html", mode="edit", e=e, subjects=subjects)

    bad = _csrf_or_flash("admin.admin_exam_edit", exam_id=exam_id)
    if bad:
        return bad

    name = (request.form.get("name") or "").strip()
    subject_id = (request.form.get("subject_id") or "").strip()
    exam_date = (request.form.get("exam_date") or "").strip()
    description = (request.form.get("description") or "").strip()

    if not name or not subject_id.isdigit():
        flash("Vui lòng nhập Tên kỳ thi và chọn Môn học.", "danger")
        return redirect(url_for("admin.admin_exam_edit", exam_id=exam_id))

    subject = Subject.query.get(int(subject_id))
    if not subject:
        flash("Môn học không tồn tại.", "danger")
        return redirect(url_for("admin.admin_exam_edit", exam_id=exam_id))

    dt = None
    if exam_date:
        try:
            dt = datetime.strptime(exam_date, "%Y-%m-%d").date()
        except Exception:
            flash("Ngày thi không đúng định dạng (YYYY-MM-DD).", "danger")
            return redirect(url_for("admin.admin_exam_edit", exam_id=exam_id))

    e.name = name
    e.subject_id = subject.id
    e.exam_date = dt
    e.description = description or None
    db.session.commit()

    flash("Đã cập nhật kỳ thi!", "success")
    return redirect(url_for("admin.admin_exams"))


@admin_bp.post("/exams/<int:exam_id>/delete")
@login_required
@roles_required("admin")
def admin_exam_delete(exam_id: int):
    e = Exam.query.get(exam_id)
    if not e:
        abort(404)

    bad = _csrf_or_flash("admin.admin_exams")
    if bad:
        return bad

    db.session.delete(e)
    db.session.commit()
    flash("Đã xoá kỳ thi!", "success")
    return redirect(url_for("admin.admin_exams"))


# =========================================================
# GRADES (LIST + CRUD)
# =========================================================
@admin_bp.get("/grades")
@login_required
@roles_required("admin")
def admin_grades():
    student_q = (request.args.get("student_q") or "").strip()
    subject_id = (request.args.get("subject_id") or "").strip()
    exam_id = (request.args.get("exam_id") or "").strip()

    query = Grade.query

    if student_q:
        query = query.join(Student).filter(
            (Student.ma_sv.ilike(f"%{student_q}%")) |
            (Student.ten_sv.ilike(f"%{student_q}%")) |
            (Student.email.ilike(f"%{student_q}%"))
        )

    if subject_id.isdigit():
        query = query.filter(Grade.subject_id == int(subject_id))

    if exam_id.isdigit():
        query = query.filter(Grade.exam_id == int(exam_id))

    grades = query.order_by(Grade.id.desc()).all()
    subjects = Subject.query.order_by(Subject.id.desc()).all()
    exams = Exam.query.order_by(Exam.id.desc()).all()

    return render_template(
        "admin/grades.html",
        grades=grades,
        subjects=subjects,
        exams=exams,
        student_q=student_q,
        subject_id=subject_id,
        exam_id=exam_id,
    )


@admin_bp.route("/grades/new", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def admin_grade_create():
    if request.method == "GET":
        students = Student.query.filter(Student.role == "student").order_by(Student.id.desc()).all()
        subjects = Subject.query.order_by(Subject.name.asc()).all()
        exams = Exam.query.order_by(Exam.id.desc()).all()
        return render_template("admin/grade_form.html", mode="create", g=None, students=students, subjects=subjects, exams=exams)

    bad = _csrf_or_flash("admin.admin_grade_create")
    if bad:
        return bad

    student_id = request.form.get("student_id", "")
    subject_id = request.form.get("subject_id", "")
    exam_id = request.form.get("exam_id", "")
    score_raw = (request.form.get("score") or "").strip()
    note = (request.form.get("note") or "").strip()

    if not (student_id.isdigit() and subject_id.isdigit() and exam_id.isdigit()):
        flash("Vui lòng chọn Sinh viên / Môn / Kỳ thi hợp lệ.", "danger")
        return redirect(url_for("admin.admin_grade_create"))

    try:
        score = float(score_raw)
    except Exception:
        flash("Điểm không hợp lệ (phải là số).", "danger")
        return redirect(url_for("admin.admin_grade_create"))

    if score < 0 or score > 10:
        flash("Điểm phải nằm trong khoảng 0 - 10.", "danger")
        return redirect(url_for("admin.admin_grade_create"))

    exists = Grade.query.filter_by(
        student_id=int(student_id),
        subject_id=int(subject_id),
        exam_id=int(exam_id),
    ).first()
    if exists:
        flash("Điểm đã tồn tại cho SV - Môn - Kỳ thi này.", "danger")
        return redirect(url_for("admin.admin_grade_create"))

    g = Grade(
        student_id=int(student_id),
        subject_id=int(subject_id),
        exam_id=int(exam_id),
        score=score,
        note=note or None,
    )
    db.session.add(g)
    db.session.commit()
    flash("Đã thêm điểm!", "success")
    return redirect(url_for("admin.admin_grades"))


@admin_bp.route("/grades/<int:grade_id>/edit", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def admin_grade_edit(grade_id: int):
    g = Grade.query.get(grade_id)
    if not g:
        abort(404)

    if request.method == "GET":
        students = Student.query.filter(Student.role == "student").order_by(Student.id.desc()).all()
        subjects = Subject.query.order_by(Subject.name.asc()).all()
        exams = Exam.query.order_by(Exam.id.desc()).all()
        return render_template("admin/grade_form.html", mode="edit", g=g, students=students, subjects=subjects, exams=exams)

    bad = _csrf_or_flash("admin.admin_grade_edit", grade_id=grade_id)
    if bad:
        return bad

    student_id = request.form.get("student_id", "")
    subject_id = request.form.get("subject_id", "")
    exam_id = request.form.get("exam_id", "")
    score_raw = (request.form.get("score") or "").strip()
    note = (request.form.get("note") or "").strip()

    if not (student_id.isdigit() and subject_id.isdigit() and exam_id.isdigit()):
        flash("Vui lòng chọn Sinh viên / Môn / Kỳ thi hợp lệ.", "danger")
        return redirect(url_for("admin.admin_grade_edit", grade_id=grade_id))

    try:
        score = float(score_raw)
    except Exception:
        flash("Điểm không hợp lệ (phải là số).", "danger")
        return redirect(url_for("admin.admin_grade_edit", grade_id=grade_id))

    if score < 0 or score > 10:
        flash("Điểm phải nằm trong khoảng 0 - 10.", "danger")
        return redirect(url_for("admin.admin_grade_edit", grade_id=grade_id))

    exists = Grade.query.filter(
        Grade.student_id == int(student_id),
        Grade.subject_id == int(subject_id),
        Grade.exam_id == int(exam_id),
        Grade.id != g.id
    ).first()
    if exists:
        flash("Bị trùng: đã có điểm cho SV - Môn - Kỳ thi này.", "danger")
        return redirect(url_for("admin.admin_grade_edit", grade_id=grade_id))

    g.student_id = int(student_id)
    g.subject_id = int(subject_id)
    g.exam_id = int(exam_id)
    g.score = score
    g.note = note or None
    db.session.commit()

    flash("Đã cập nhật điểm!", "success")
    return redirect(url_for("admin.admin_grades"))


@admin_bp.post("/grades/<int:grade_id>/delete")
@login_required
@roles_required("admin")
def admin_grade_delete(grade_id: int):
    g = Grade.query.get(grade_id)
    if not g:
        abort(404)

    bad = _csrf_or_flash("admin.admin_grades")
    if bad:
        return bad

    db.session.delete(g)
    db.session.commit()
    flash("Đã xoá điểm!", "success")
    return redirect(url_for("admin.admin_grades"))


# =========================================================
# LESSONS (LIST + CRUD)
# =========================================================
@admin_bp.get("/lessons")
@login_required
@roles_required("admin")
def admin_lessons():
    subject_id = request.args.get("subject_id", type=int)
    subjects = Subject.query.order_by(Subject.name.asc()).all()

    q = Lesson.query
    if subject_id:
        q = q.filter(Lesson.subject_id == subject_id)

    lessons = q.order_by(Lesson.subject_id.asc(), Lesson.order_no.asc(), Lesson.id.asc()).all()

    return render_template("admin/lessons.html", lessons=lessons, subjects=subjects, subject_id=subject_id)


@admin_bp.route("/lessons/new", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def admin_lessons_create():
    subjects = Subject.query.order_by(Subject.name.asc()).all()
    if not subjects:
        flash("Chưa có môn học. Hãy tạo môn học trước!", "warning")
        return redirect(url_for("admin.admin_subjects"))

    if request.method == "GET":
        return render_template("admin/lesson_form.html", mode="create", subjects=subjects, lesson=None)

    bad = _csrf_or_flash("admin.admin_lessons_create")
    if bad:
        return bad

    subject_id = request.form.get("subject_id", "")
    title = (request.form.get("title") or "").strip()
    content = (request.form.get("content") or "").strip()
    video_url = (request.form.get("video_url") or "").strip()
    order_no = request.form.get("order_no", type=int) or 1

    if not subject_id.isdigit():
        flash("Vui lòng chọn môn học hợp lệ.", "danger")
        return redirect(url_for("admin.admin_lessons_create"))

    if not title:
        flash("Tiêu đề bài học không được trống.", "danger")
        return redirect(url_for("admin.admin_lessons_create"))

    lesson = Lesson(
        subject_id=int(subject_id),
        title=title,
        content=content or None,
        video_url=video_url or None,
        order_no=order_no
    )
    db.session.add(lesson)
    db.session.commit()
    flash("Đã thêm bài học!", "success")
    return redirect(url_for("admin.admin_lessons"))


@admin_bp.route("/lessons/<int:lesson_id>/edit", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def admin_lessons_edit(lesson_id: int):
    lesson = Lesson.query.get(lesson_id)
    if not lesson:
        abort(404)

    subjects = Subject.query.order_by(Subject.name.asc()).all()

    if request.method == "GET":
        return render_template("admin/lesson_form.html", mode="edit", subjects=subjects, lesson=lesson)

    bad = _csrf_or_flash("admin.admin_lessons_edit", lesson_id=lesson_id)
    if bad:
        return bad

    subject_id = request.form.get("subject_id", "")
    title = (request.form.get("title") or "").strip()
    content = (request.form.get("content") or "").strip()
    video_url = (request.form.get("video_url") or "").strip()
    order_no = request.form.get("order_no", type=int) or 1

    if not subject_id.isdigit():
        flash("Vui lòng chọn môn học hợp lệ.", "danger")
        return redirect(url_for("admin.admin_lessons_edit", lesson_id=lesson_id))

    if not title:
        flash("Tiêu đề bài học không được trống.", "danger")
        return redirect(url_for("admin.admin_lessons_edit", lesson_id=lesson_id))

    lesson.subject_id = int(subject_id)
    lesson.title = title
    lesson.content = content or None
    lesson.video_url = video_url or None
    lesson.order_no = order_no

    db.session.commit()
    flash("Đã cập nhật bài học!", "success")
    return redirect(url_for("admin.admin_lessons"))


@admin_bp.post("/lessons/<int:lesson_id>/delete")
@login_required
@roles_required("admin")
def admin_lessons_delete(lesson_id: int):
    lesson = Lesson.query.get(lesson_id)
    if not lesson:
        abort(404)

    bad = _csrf_or_flash("admin.admin_lessons")
    if bad:
        return bad

    db.session.delete(lesson)
    db.session.commit()
    flash("Đã xoá bài học!", "success")
    return redirect(url_for("admin.admin_lessons"))


# =========================================================
# STUDENTS (ADMIN) + Import/Export/Report
# =========================================================
@admin_bp.get("/students")
@login_required
@roles_required("admin")
def admin_students_index():
    q_ma = (request.args.get("ma_sv") or "").strip()
    q_ten = (request.args.get("ten_sv") or "").strip()

    try:
        page = int(request.args.get("page", 1))
    except ValueError:
        page = 1

    try:
        per_page = int(request.args.get("per_page", 5))
    except ValueError:
        per_page = 5
    per_page = 5 if per_page not in (5, 10, 20) else per_page

    query = Student.query.filter(Student.role == "student")
    if q_ma:
        query = query.filter(Student.ma_sv.ilike(f"%{q_ma}%"))
    if q_ten:
        query = query.filter(Student.ten_sv.ilike(f"%{q_ten}%"))

    pagination = query.order_by(Student.id.asc()).paginate(page=page, per_page=per_page, error_out=False)

    return render_template(
        "admin/students.html",
        students=pagination.items,
        pagination=pagination,
        ma_sv=q_ma,
        ten_sv=q_ten,
        per_page=per_page,
    )


@admin_bp.route("/students/new", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def admin_students_create():
    if request.method == "GET":
        return render_template("admin/student_form.html", mode="create", student=None)

    bad = _csrf_or_flash("admin.admin_students_create")
    if bad:
        return bad

    ma_sv = (request.form.get("ma_sv") or "").strip()
    ten_sv = (request.form.get("ten_sv") or "").strip()
    email = (request.form.get("email") or "").strip().lower()
    mat_khau = (request.form.get("mat_khau") or "").strip()
    gioi_tinh = (request.form.get("gioi_tinh") or "").strip()
    dia_chi = (request.form.get("dia_chi") or "").strip()

    if not ma_sv or not ten_sv or not email or not mat_khau:
        flash("Vui lòng nhập đủ: Mã SV, Tên, Email, Mật khẩu.", "danger")
        return redirect(url_for("admin.admin_students_create"))

    if not re.fullmatch(r"[A-Za-z0-9_-]+", ma_sv):
        flash("Mã SV chỉ gồm chữ/số/_/- (không ký tự đặc biệt).", "danger")
        return redirect(url_for("admin.admin_students_create"))

    if Student.query.filter_by(ma_sv=ma_sv).first():
        flash("Mã SV đã tồn tại.", "danger")
        return redirect(url_for("admin.admin_students_create"))

    if Student.query.filter_by(email=email).first():
        flash("Email đã tồn tại.", "danger")
        return redirect(url_for("admin.admin_students_create"))

    hashed = bcrypt.generate_password_hash(mat_khau).decode("utf-8")
    sv = Student(
        ma_sv=ma_sv,
        ten_sv=ten_sv,
        email=email,
        mat_khau=hashed,
        gioi_tinh=gioi_tinh or None,
        dia_chi=dia_chi or None,
        role="student",
    )

    db.session.add(sv)
    db.session.commit()
    flash("Thêm sinh viên thành công!", "success")
    return redirect(url_for("admin.admin_students_index"))


@admin_bp.route("/students/<int:student_id>/edit", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def admin_students_edit(student_id: int):
    sv = Student.query.get(student_id)
    if not sv:
        abort(404)

    if request.method == "GET":
        return render_template("admin/student_form.html", mode="edit", student=sv)

    bad = _csrf_or_flash("admin.admin_students_edit", student_id=student_id)
    if bad:
        return bad

    ma_sv = (request.form.get("ma_sv") or "").strip()
    ten_sv = (request.form.get("ten_sv") or "").strip()
    email = (request.form.get("email") or "").strip().lower()
    mat_khau = (request.form.get("mat_khau") or "").strip()  # optional
    gioi_tinh = (request.form.get("gioi_tinh") or "").strip()
    dia_chi = (request.form.get("dia_chi") or "").strip()

    if not ma_sv or not ten_sv or not email:
        flash("Vui lòng nhập đủ: Mã SV, Tên, Email.", "danger")
        return redirect(url_for("admin.admin_students_edit", student_id=student_id))

    if not re.fullmatch(r"[A-Za-z0-9_-]+", ma_sv):
        flash("Mã SV chỉ gồm chữ/số/_/- (không ký tự đặc biệt).", "danger")
        return redirect(url_for("admin.admin_students_edit", student_id=student_id))

    if Student.query.filter(Student.ma_sv == ma_sv, Student.id != sv.id).first():
        flash("Mã SV đã tồn tại.", "danger")
        return redirect(url_for("admin.admin_students_edit", student_id=student_id))

    if Student.query.filter(Student.email == email, Student.id != sv.id).first():
        flash("Email đã tồn tại.", "danger")
        return redirect(url_for("admin.admin_students_edit", student_id=student_id))

    sv.ma_sv = ma_sv
    sv.ten_sv = ten_sv
    sv.email = email
    sv.gioi_tinh = gioi_tinh or None
    sv.dia_chi = dia_chi or None

    if mat_khau:
        sv.mat_khau = bcrypt.generate_password_hash(mat_khau).decode("utf-8")

    db.session.commit()
    flash("Cập nhật sinh viên thành công!", "success")
    return redirect(url_for("admin.admin_students_index"))


@admin_bp.post("/students/<int:student_id>/delete")
@login_required
@roles_required("admin")
def admin_students_delete(student_id: int):
    sv = Student.query.get(student_id)
    if not sv:
        abort(404)

    bad = _csrf_or_flash("admin.admin_students_index")
    if bad:
        return bad

    if sv.avatar:
        old_path = os.path.join(_upload_dir(), sv.avatar)
        if os.path.exists(old_path):
            try:
                os.remove(old_path)
            except Exception:
                pass

    db.session.delete(sv)
    db.session.commit()
    flash("Xóa sinh viên thành công!", "success")
    return redirect(url_for("admin.admin_students_index"))


@admin_bp.post("/students/<int:student_id>/avatar")
@login_required
@roles_required("admin")
def admin_student_upload_avatar(student_id: int):
    sv = Student.query.get(student_id)
    if not sv:
        abort(404)

    bad = _csrf_or_flash("admin.admin_students_index")
    if bad:
        return bad

    f = request.files.get("avatar")
    if not f or f.filename == "":
        flash("Chưa chọn ảnh.", "danger")
        return redirect(url_for("admin.admin_students_index"))

    filename = secure_filename(f.filename)
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_EXTS:
        flash("Chỉ nhận ảnh PNG/JPG/WEBP.", "danger")
        return redirect(url_for("admin.admin_students_index"))

    if sv.avatar:
        old_path = os.path.join(_upload_dir(), sv.avatar)
        if os.path.exists(old_path):
            try:
                os.remove(old_path)
            except Exception:
                pass

    new_name = f"{uuid.uuid4().hex}{ext}"
    f.save(os.path.join(_upload_dir(), new_name))

    sv.avatar = new_name
    db.session.commit()
    flash("Cập nhật avatar thành công!", "success")
    return redirect(url_for("admin.admin_students_index"))


@admin_bp.get("/students/export")
@login_required
@roles_required("admin")
def admin_students_export():
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(EXPORT_HEADERS)

    for s in Student.query.filter(Student.role == "student").order_by(Student.id.asc()).all():
        new_plain_password = _random_password(10)
        ws.append([s.ma_sv, s.ten_sv, s.email, new_plain_password, s.gioi_tinh or "", s.dia_chi or ""])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="students.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@admin_bp.route("/students/import", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def admin_students_import():
    if request.method == "GET":
        return render_template("admin/import.html")

    bad = _csrf_or_flash("admin.admin_students_import")
    if bad:
        return bad

    f = request.files.get("file")
    if not f or f.filename == "":
        flash("Vui lòng chọn file Excel (.xlsx)", "danger")
        return redirect(url_for("admin.admin_students_import"))

    wb = load_workbook(f)
    ws = wb.active

    header_cells = [(_norm_cell(c.value)) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_index = {name: idx for idx, name in enumerate(header_cells) if name}

    missing = [h for h in EXPORT_HEADERS if h not in col_index]
    if missing:
        flash(f"File thiếu cột: {', '.join(missing)}", "danger")
        return redirect(url_for("admin.admin_students_import"))

    added, skipped = 0, 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        ma_sv = _norm_cell(row[col_index["Mã SV"]] if col_index["Mã SV"] < len(row) else None)
        ten_sv = _norm_cell(row[col_index["Tên"]] if col_index["Tên"] < len(row) else None)
        email = _norm_cell(row[col_index["Email"]] if col_index["Email"] < len(row) else None).lower()
        mat_khau_plain = _norm_cell(row[col_index["Mật khẩu"]] if col_index["Mật khẩu"] < len(row) else None)
        gioi_tinh = _norm_cell(row[col_index["Giới tính"]] if col_index["Giới tính"] < len(row) else None)
        dia_chi = _norm_cell(row[col_index["Địa chỉ"]] if col_index["Địa chỉ"] < len(row) else None)

        if not ma_sv or not ten_sv or not email:
            skipped += 1
            continue

        if not re.fullmatch(r"[A-Za-z0-9_-]+", ma_sv):
            skipped += 1
            continue

        if not mat_khau_plain:
            mat_khau_plain = _random_password(10)

        hashed = bcrypt.generate_password_hash(mat_khau_plain).decode("utf-8")

        sv = Student.query.filter_by(ma_sv=ma_sv).first()
        if sv is None:
            sv_by_email = Student.query.filter_by(email=email).first()
            if sv_by_email is not None:
                sv = sv_by_email
            else:
                sv = Student(ma_sv=ma_sv)
                db.session.add(sv)
                added += 1

        sv.ma_sv = ma_sv
        sv.ten_sv = ten_sv
        sv.email = email
        sv.mat_khau = hashed
        sv.gioi_tinh = gioi_tinh or None
        sv.dia_chi = dia_chi or None
        sv.role = "student"  # ✅ ép role student

    db.session.commit()
    flash(f"Import xong: thêm {added} sinh viên, bỏ qua {skipped} dòng không hợp lệ.", "success")
    return redirect(url_for("admin.admin_students_index"))


@admin_bp.get("/students/report")
@login_required
@roles_required("admin")
def admin_students_report():
    q = Student.query.filter(Student.role == "student")
    total = q.count()
    nam = q.filter(Student.gioi_tinh == "Nam").count()
    nu = q.filter(Student.gioi_tinh == "Nữ").count()
    khac = total - nam - nu
    return render_template("admin/report.html", total=total, nam=nam, nu=nu, khac=khac)


# =========================================================
# TEACHERS (gán lớp cho giáo viên)
# =========================================================
@admin_bp.get("/teachers")
@login_required
@roles_required("admin")
def admin_teachers():
    teachers = Student.query.filter_by(role="manager").order_by(Student.id.desc()).all()
    classes = Class.query.order_by(Class.name.asc()).all()
    return render_template("admin/teachers.html", teachers=teachers, classes=classes)


@admin_bp.post("/teachers/<int:teacher_id>/classes")
@login_required
@roles_required("admin")
def admin_teacher_update_classes(teacher_id: int):
    teacher = Student.query.get(teacher_id)
    if not teacher or teacher.role != "manager":
        abort(404)

    bad = _csrf_or_flash("admin.admin_teachers")
    if bad:
        return bad

    class_ids = request.form.getlist("class_ids")
    class_ids = [int(x) for x in class_ids if str(x).isdigit()]

    teacher.teaching_classes = Class.query.filter(Class.id.in_(class_ids)).all() if class_ids else []
    db.session.commit()

    flash("Đã cập nhật lớp phụ trách cho giáo viên!", "success")
    return redirect(url_for("admin.admin_teachers"))
