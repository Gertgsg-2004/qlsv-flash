from __future__ import annotations

import os
import re
import smtplib
import secrets
import string
import uuid
from datetime import datetime, timedelta
from io import BytesIO
from typing import Optional
from collections import Counter
from auth import auth_bp


from flask import Flask, render_template, redirect, url_for, request, flash, abort, send_file
from flask_wtf import FlaskForm
from flask_wtf.csrf import generate_csrf
from markupsafe import Markup
from sqlalchemy.exc import IntegrityError
from werkzeug.utils import secure_filename
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from itsdangerous import URLSafeTimedSerializer
from openpyxl import Workbook, load_workbook
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from wtforms import StringField, PasswordField, RadioField, TextAreaField, SubmitField, SelectField
from wtforms.validators import DataRequired, Email, Optional as Opt, Length, ValidationError

from extensions import db, bcrypt, csrf  # ✅ chỉ dùng extensions
from account_routes import account_bp
from notifications_routes import noti_bp




# -------------------------
# App setup
# -------------------------
app = Flask(__name__)
app.config["SECRET_KEY"] = "dev-secret-key-change-me"

# --- Email SMTP config ---
app.config["MAIL_HOST"] = "smtp.gmail.com"
app.config["MAIL_PORT"] = 587
app.config["MAIL_USE_TLS"] = True
app.config["MAIL_USERNAME"] = "your_email@gmail.com"
app.config["MAIL_PASSWORD"] = "your_app_password"
app.config["MAIL_FROM"] = "QLSV <your_email@gmail.com>"

# DB in instance/
os.makedirs(app.instance_path, exist_ok=True)
db_path = os.path.join(app.instance_path, "qlsv.sqlite3")
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + db_path
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# ✅ init extensions
db.init_app(app)
bcrypt.init_app(app)
csrf.init_app(app)

# Upload avatar
UPLOAD_DIR = os.path.join(app.root_path, "static", "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)
ALLOWED_EXTS = {".png", ".jpg", ".jpeg", ".webp"}

# Login
login_manager = LoginManager(app)
login_manager.login_view = "login"

# Reset password token serializer
serializer = URLSafeTimedSerializer(app.config["SECRET_KEY"])

# ✅ IMPORT MODEL 1 LẦN DUY NHẤT
from models import Notification, Student  # Student nằm ở models.py

# ✅ tạo bảng + seed admin (chạy 1 lần)
with app.app_context():
    db.create_all()

    admin_email = "admin@gmail.com"
    if Student.query.filter_by(email=admin_email).first() is None:
        hashed = bcrypt.generate_password_hash("123456").decode("utf-8")
        db.session.add(
            Student(ma_sv="ADMIN", ten_sv="Admin", email=admin_email, mat_khau=hashed, role="admin")
        )
        db.session.commit()

# ✅ Import blueprints SAU CÙNG (tránh vòng import)
from admin_routes import _csrf_or_flash, admin_bp
from manager_routes import manager_bp
from student_routes import student_bp

app.register_blueprint(admin_bp)
app.register_blueprint(manager_bp)
app.register_blueprint(student_bp)
app.register_blueprint(auth_bp)
app.register_blueprint(account_bp)
app.register_blueprint(noti_bp)


# -------------------------
# CSRF helper (để form delete/upload không in ra code)
# -------------------------
@app.context_processor
def inject_csrf():
    def form_csrf_token():
        return Markup(f'<input type="hidden" name="csrf_token" value="{generate_csrf()}">')
    return dict(form_csrf_token=form_csrf_token)

# -------------------------
# Email helpers
# -------------------------
def _mail_is_configured() -> bool:
    u = (app.config.get("MAIL_USERNAME") or "").strip()
    p = (app.config.get("MAIL_PASSWORD") or "").strip()
    return u and p and "your_email" not in u and "your_app_password" not in p

def send_email(to_email: str, subject: str, html: str) -> bool:
    """Gửi email HTML qua SMTP. True nếu gửi thật; False nếu DEMO."""
    if not _mail_is_configured():
        return False

    host = app.config["MAIL_HOST"]
    port = app.config["MAIL_PORT"]
    username = app.config["MAIL_USERNAME"]
    password = app.config["MAIL_PASSWORD"]
    from_email = app.config["MAIL_FROM"]

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = from_email
    msg["To"] = to_email
    msg.attach(MIMEText(html, "html", "utf-8"))

    with smtplib.SMTP(host, port) as server:
        if app.config.get("MAIL_USE_TLS"):
            server.starttls()
        server.login(username, password)
        server.sendmail(username, to_email, msg.as_string())

    return True

# -------------------------
# Reset token helpers
# -------------------------
def generate_reset_token(email: str) -> str:
    return serializer.dumps(email, salt="reset-password")

def verify_reset_token(token: str, max_age_seconds: int = 15 * 60) -> str | None:
    try:
        email = serializer.loads(token, salt="reset-password", max_age=max_age_seconds)
        return (email or "").strip().lower()
    except Exception:
        return None

def generate_random_password(length: int = 10) -> str:
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))

def _norm_cell(x) -> str:
    if x is None:
        return ""
    return str(x).strip()

# -------------------------
# Forms
# -------------------------
class StudentForm(FlaskForm):
    ma_sv = StringField("Mã sinh viên", validators=[DataRequired(message="Mã SV không được bỏ trống.")])
    ten_sv = StringField("Họ và tên", validators=[DataRequired(message="Tên không được bỏ trống.")])

    email = StringField(
        "Email",
        validators=[DataRequired(message="Email không được bỏ trống."), Email(message="Email không đúng định dạng.")]
    )

    mat_khau = PasswordField("Mật khẩu")

    gioi_tinh = RadioField(
        "Giới tính",
        choices=[("Nam", "Nam"), ("Nữ", "Nữ"), ("Khác", "Khác")],
        validators=[Opt()],
    )
    dia_chi = TextAreaField("Địa chỉ", validators=[Opt()])
    submit = SubmitField("Lưu")

    editing_id: Optional[int] = None

    def validate_ma_sv(self, field):
        ma = (field.data or "").strip()
        if not re.fullmatch(r"[A-Za-z0-9_-]+", ma):
            raise ValidationError("Mã SV chỉ gồm chữ/số/_/- (không ký tự đặc biệt).")

        q = Student.query.filter_by(ma_sv=ma)
        if self.editing_id is not None:
            q = q.filter(Student.id != self.editing_id)
        if q.first() is not None:
            raise ValidationError("Mã SV đã tồn tại.")

    def validate_email(self, field):
        em = (field.data or "").strip().lower()
        q = Student.query.filter(Student.email == em)
        if self.editing_id is not None:
            q = q.filter(Student.id != self.editing_id)
        if q.first() is not None:
            raise ValidationError("Email đã tồn tại.")

class LoginForm(FlaskForm):
    identifier = StringField("Mã SV hoặc Email", validators=[DataRequired(message="Nhập mã SV hoặc email.")])
    password = PasswordField("Mật khẩu", validators=[DataRequired(message="Nhập mật khẩu.")])
    submit = SubmitField("Đăng nhập")

class ForgotForm(FlaskForm):
    email = StringField("Email", validators=[DataRequired(), Email()])
    method = SelectField(
        "Chọn cách khôi phục",
        choices=[
            ("newpass", "Cách 1: Cấp mật khẩu mới và gửi qua email"),
            ("link", "Cách 2: Gửi link đặt lại mật khẩu"),
        ],
        validators=[DataRequired()],
    )
    submit = SubmitField("Thực hiện")

class ResetForm(FlaskForm):
    password = PasswordField("Mật khẩu mới", validators=[DataRequired(), Length(min=6, message="Tối thiểu 6 ký tự")])
    submit = SubmitField("Đổi mật khẩu")

# -------------------------
# Login user_loader
# -------------------------
@login_manager.user_loader
def load_user(user_id: str):
    try:
        return Student.query.get(int(user_id))
    except Exception:
        return None

# -------------------------
# Login lock helpers
# -------------------------
def is_locked(user: Student) -> bool:
    return bool(user.lock_until and user.lock_until > datetime.utcnow())

def record_failed_login(user: Student):
    now = datetime.utcnow()
    if user.last_failed_at and (now - user.last_failed_at) > timedelta(seconds=60):
        user.failed_attempts = 0

    user.failed_attempts += 1
    user.last_failed_at = now

    if user.failed_attempts >= 3:
        user.lock_until = now + timedelta(seconds=60)
        user.failed_attempts = 0

    db.session.commit()

def clear_failed_login(user: Student):
    user.failed_attempts = 0
    user.last_failed_at = None
    user.lock_until = None
    db.session.commit()

# -------------------------
# Routes - Auth
# -------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        role = (getattr(current_user, "role", "student") or "student").lower()
        if role == "admin":
            return redirect(url_for("admin.admin_home"))
        if role == "manager":
            return redirect(url_for("manager.manager_home"))
        return redirect(url_for("student.student_home"))

    form = LoginForm()

    if form.validate_on_submit():
        ident = (form.identifier.data or "").strip()
        pw = form.password.data

        user = Student.query.filter(
            (Student.ma_sv == ident) | (Student.email == ident.lower())
        ).first()

        if not user:
            flash("Sai tài khoản hoặc mật khẩu.", "danger")
            return render_template("login.html", form=form)

        if is_locked(user):
            remain = int((user.lock_until - datetime.utcnow()).total_seconds())
            flash(f"Tài khoản đang bị khóa. Thử lại sau {remain}s.", "danger")
            return render_template("login.html", form=form)

        if not user.check_password(pw):
            record_failed_login(user)
            flash("Sai tài khoản hoặc mật khẩu.", "danger")
            return render_template("login.html", form=form)

        clear_failed_login(user)
        login_user(user)
        flash("Đăng nhập thành công!", "success")

        role = (getattr(user, "role", "student") or "student").lower()
        if role == "admin":
            return redirect(url_for("admin.admin_home"))
        elif role == "manager":
            return redirect(url_for("manager.manager_home"))
        else:
            return redirect(url_for("student.student_home"))

    return render_template("login.html", form=form)

@app.get("/logout")
@login_required
def logout():
    logout_user()
    flash("Đã đăng xuất.", "success")
    return redirect(url_for("login"))

@app.get("/make-admin/<email>")
def make_admin(email):
    u = Student.query.filter_by(email=email).first()
    if not u:
        return "not found"
    u.role = "admin"
    db.session.commit()
    return "ok"

@app.route("/register", methods=["GET", "POST"])
def register():
    if current_user.is_authenticated:
        role = (getattr(current_user, "role", "student") or "student").lower()
        if role == "admin":
            return redirect(url_for("admin.admin_home"))
        if role == "manager":
            return redirect(url_for("manager.manager_home"))
        return redirect(url_for("student.student_home"))

    form = StudentForm()
    form.mat_khau.validators = [
        DataRequired(message="Mật khẩu không được bỏ trống."),
        Length(min=6, message="Mật khẩu tối thiểu 6 ký tự.")
    ]

    if form.validate_on_submit():
        hashed = bcrypt.generate_password_hash(form.mat_khau.data).decode("utf-8")
        sv = Student(
            ma_sv=form.ma_sv.data.strip(),
            ten_sv=form.ten_sv.data.strip(),
            email=form.email.data.strip().lower(),
            mat_khau=hashed,
            gioi_tinh=form.gioi_tinh.data or None,
            dia_chi=form.dia_chi.data.strip() if form.dia_chi.data else None,
        )

        try:
            db.session.add(sv)
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash("Email hoặc Mã sinh viên đã tồn tại!", "danger")
            return render_template("register.html", form=form)

        flash("Đăng ký thành công! Bạn có thể đăng nhập.", "success")
        return redirect(url_for("login"))

    return render_template("register.html", form=form)

@app.route("/forgot", methods=["GET", "POST"])
def forgot():
    form = ForgotForm()
    demo_new_password = None
    demo_reset_link = None

    if form.validate_on_submit():
        email = form.email.data.strip().lower()
        method = form.method.data

        user = Student.query.filter_by(email=email).first()
        if not user:
            flash("Email không tồn tại trong hệ thống.", "danger")
            return render_template("forgot.html", form=form)

        if method == "newpass":
            new_pass = generate_random_password(10)
            user.mat_khau = bcrypt.generate_password_hash(new_pass).decode("utf-8")
            db.session.commit()

            ok = send_email(
                to_email=email,
                subject="QLSV - Mật khẩu mới",
                html=f"""
                <p>Chào bạn,</p>
                <p>Hệ thống đã cấp mật khẩu mới cho bạn:</p>
                <h3>{new_pass}</h3>
                <p>Vui lòng đăng nhập và đổi mật khẩu sau khi vào hệ thống.</p>
                """,
            )

            if ok:
                flash("Đã gửi mật khẩu mới vào email.", "success")
            else:
                flash("SMTP chưa cấu hình => hiển thị mật khẩu mới (DEMO).", "warning")
                demo_new_password = new_pass

            return render_template("forgot.html", form=form, demo_new_password=demo_new_password)

        token = generate_reset_token(email)
        link = url_for("reset_password", token=token, _external=True)

        ok = send_email(
            to_email=email,
            subject="QLSV - Link đặt lại mật khẩu",
            html=f"""
            <p>Chào bạn,</p>
            <p>Bấm link sau để đặt lại mật khẩu (hết hạn sau 15 phút):</p>
            <p><a href="{link}">{link}</a></p>
            """,
        )

        if ok:
            flash("Đã gửi link đặt lại mật khẩu vào email.", "success")
        else:
            flash("SMTP chưa cấu hình => hiển thị link reset (DEMO).", "warning")
            demo_reset_link = link

        return render_template("forgot.html", form=form, reset_link=demo_reset_link)

    return render_template("forgot.html", form=form)

@app.route("/reset/<token>", methods=["GET", "POST"])
def reset_password(token: str):
    email = verify_reset_token(token)
    if not email:
        flash("Link reset không hợp lệ hoặc đã hết hạn.", "danger")
        return redirect(url_for("login"))

    user = Student.query.filter_by(email=email).first()
    if not user:
        flash("Tài khoản không tồn tại.", "danger")
        return redirect(url_for("login"))

    form = ResetForm()
    if form.validate_on_submit():
        user.mat_khau = bcrypt.generate_password_hash(form.password.data).decode("utf-8")
        db.session.commit()
        flash("Đổi mật khẩu thành công. Hãy đăng nhập lại.", "success")
        return redirect(url_for("login"))

    return render_template("reset.html", form=form)

# -------------------------
# Root / Students redirect
# -------------------------
@app.get("/")
def root():
    return redirect(url_for("students_index"))

@app.get("/students")
@login_required
def students_index():
    role = (getattr(current_user, "role", "student") or "student").lower()
    if role == "admin":
        return redirect(url_for("admin.admin_students_index"))
    if role == "manager":
        return redirect(url_for("manager.manager_home"))
    return redirect(url_for("student.student_home"))

# -------------------------
# Students CRUD + Avatar + Import/Export + Report (GIỮ NGUYÊN)
# (Bạn đang dùng admin/student qua blueprint; phần này bạn có thể giữ để debug,
#  nhưng nó sẽ không chạy vì /students đã redirect theo role)
# -------------------------
@app.route("/students/new", methods=["GET", "POST"])
@login_required
def students_create():
    form = StudentForm()
    form.mat_khau.validators = [DataRequired(message="Mật khẩu không được bỏ trống."), Length(min=6)]

    if form.validate_on_submit():
        hashed = bcrypt.generate_password_hash(form.mat_khau.data).decode("utf-8")
        sv = Student(
            ma_sv=form.ma_sv.data.strip(),
            ten_sv=form.ten_sv.data.strip(),
            email=form.email.data.strip().lower(),
            mat_khau=hashed,
            gioi_tinh=form.gioi_tinh.data or None,
            dia_chi=form.dia_chi.data.strip() if form.dia_chi.data else None,
        )

        try:
            db.session.add(sv)
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash("Mã SV hoặc Email đã tồn tại!", "danger")
            return render_template("form.html", form=form, mode="create")

        flash("Thêm sinh viên thành công!", "success")
        return redirect(url_for("students_index"))

    return render_template("form.html", form=form, mode="create")

@app.route("/students/<int:student_id>/edit", methods=["GET", "POST"])
@login_required
def students_edit(student_id: int):
    sv = Student.query.get(student_id)
    if sv is None:
        abort(404)

    form = StudentForm(obj=sv)
    form.editing_id = sv.id
    form.mat_khau.validators = [Opt(), Length(min=6)]

    if request.method == "GET":
        form.gioi_tinh.data = sv.gioi_tinh or ""
        return render_template("form.html", form=form, mode="edit", student=sv)

    if form.validate_on_submit():
        sv.ma_sv = form.ma_sv.data.strip()
        sv.ten_sv = form.ten_sv.data.strip()
        sv.email = form.email.data.strip().lower()
        sv.gioi_tinh = form.gioi_tinh.data or None
        sv.dia_chi = form.dia_chi.data.strip() if form.dia_chi.data else None

        if form.mat_khau.data:
            sv.mat_khau = bcrypt.generate_password_hash(form.mat_khau.data).decode("utf-8")

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash("Mã SV hoặc Email đã tồn tại!", "danger")
            return render_template("form.html", form=form, mode="edit", student=sv)

        flash("Cập nhật sinh viên thành công!", "success")
        return redirect(url_for("students_index"))

    return render_template("form.html", form=form, mode="edit", student=sv)

@app.post("/students/<int:student_id>/delete")
@login_required
def students_delete(student_id: int):
    sv = Student.query.get(student_id)
    if sv is None:
        abort(404)

    if sv.avatar:
        old_path = os.path.join(UPLOAD_DIR, sv.avatar)
        if os.path.exists(old_path):
            try:
                os.remove(old_path)
            except Exception:
                pass

    db.session.delete(sv)
    db.session.commit()
    flash("Xóa sinh viên thành công!", "success")
    return redirect(url_for("students_index"))

@app.post("/students/<int:student_id>/avatar")
@login_required
def student_upload_avatar(student_id: int):
    sv = Student.query.get(student_id)
    if sv is None:
        abort(404)

    f = request.files.get("avatar")
    if not f or f.filename == "":
        flash("Chưa chọn ảnh", "danger")
        return redirect(url_for("students_index"))

    filename = secure_filename(f.filename)
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_EXTS:
        flash("Chỉ nhận ảnh PNG/JPG/WEBP", "danger")
        return redirect(url_for("students_index"))

    if sv.avatar:
        old_path = os.path.join(UPLOAD_DIR, sv.avatar)
        if os.path.exists(old_path):
            try:
                os.remove(old_path)
            except Exception:
                pass

    new_name = f"{uuid.uuid4().hex}{ext}"
    f.save(os.path.join(UPLOAD_DIR, new_name))
    sv.avatar = new_name
    db.session.commit()

    flash("Cập nhật avatar sinh viên thành công!", "success")
    return redirect(url_for("students_index"))

EXPORT_HEADERS = ["Mã SV", "Tên", "Email", "Mật khẩu", "Giới tính", "Địa chỉ"]

@app.get("/students/export")
@login_required
def students_export():
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(EXPORT_HEADERS)

    for s in Student.query.order_by(Student.id.asc()).all():
        ws.append([s.ma_sv, s.ten_sv, s.email, generate_random_password(10), s.gioi_tinh or "", s.dia_chi or ""])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name="students.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.route("/students/import", methods=["GET", "POST"])
@login_required
def students_import():
    if request.method == "GET":
        return render_template("import.html")

    f = request.files.get("file")
    if not f or f.filename == "":
        flash("Vui lòng chọn file Excel (.xlsx)", "danger")
        return redirect(url_for("students_import"))

    wb = load_workbook(f)
    ws = wb.active

    header_cells = [(_norm_cell(c.value)) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_index = {name: idx for idx, name in enumerate(header_cells) if name}

    missing = [h for h in EXPORT_HEADERS if h not in col_index]
    if missing:
        flash(f"File thiếu cột: {', '.join(missing)}", "danger")
        return redirect(url_for("students_import"))

    added = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        ma_sv = _norm_cell(row[col_index["Mã SV"]] if col_index["Mã SV"] < len(row) else None)
        ten_sv = _norm_cell(row[col_index["Tên"]] if col_index["Tên"] < len(row) else None)
        email = _norm_cell(row[col_index["Email"]] if col_index["Email"] < len(row) else None).lower()
        mat_khau_plain = _norm_cell(row[col_index["Mật khẩu"]] if col_index["Mật khẩu"] < len(row) else None)
        gioi_tinh = _norm_cell(row[col_index["Giới tính"]] if col_index["Giới tính"] < len(row) else None)
        dia_chi = _norm_cell(row[col_index["Địa chỉ"]] if col_index["Địa chỉ"] < len(row) else None)

        if not ma_sv or not ten_sv or not email:
            skipped += 1
            continue
        if not re.fullmatch(r"[A-Za-z0-9_-]+", ma_sv):
            skipped += 1
            continue
        if not mat_khau_plain:
            mat_khau_plain = generate_random_password(10)

        hashed = bcrypt.generate_password_hash(mat_khau_plain).decode("utf-8")

        sv = Student.query.filter_by(ma_sv=ma_sv).first()
        if sv is None:
            sv = Student(ma_sv=ma_sv)
            db.session.add(sv)
            added += 1

        sv.ma_sv = ma_sv
        sv.ten_sv = ten_sv
        sv.email = email
        sv.mat_khau = hashed
        sv.gioi_tinh = gioi_tinh or None
        sv.dia_chi = dia_chi or None

    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        flash("Import lỗi do trùng Mã SV/Email (unique). Hãy kiểm tra file.", "danger")
        return redirect(url_for("students_import"))

    flash(f"Import xong: thêm {added} sinh viên, bỏ qua {skipped} dòng không hợp lệ.", "success")
    return redirect(url_for("students_index"))

@app.get("/report")
@login_required
def report():
    total = Student.query.count()
    counts = Counter([(s.gioi_tinh or "Khác") for s in Student.query.all()])
    nam = counts.get("Nam", 0)
    nu = counts.get("Nữ", 0)
    khac = total - nam - nu
    return render_template("report.html", total=total, nam=nam, nu=nu, khac=khac)

@app.errorhandler(404)
def not_found(_):
    return "404 - Không tìm thấy", 404

if __name__ == "__main__":
    app.run(debug=True)
BASE_DIR = os.path.abspath(os.path.dirname(__file__))

UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 2 * 1024 * 1024  # 2MB
@app.context_processor
def inject_noti_count():
    if current_user.is_authenticated:
        count = Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
        return {"noti_unread_count": count}
    return {"noti_unread_count": 0}